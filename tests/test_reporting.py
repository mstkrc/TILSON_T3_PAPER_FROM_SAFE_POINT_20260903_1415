"""Phase-14 report/export validation."""
from datetime import datetime, timezone
from decimal import Decimal
import json
from pathlib import Path
from openpyxl import load_workbook
from src.paper.ledger import EntryLedgerRecord, ExitLedgerRecord
from src.report.reporting import SHEETS, ReportFilters, build_report, export_xlsx

ROOT = Path(__file__).parents[1]
NOW = datetime(2026, 1, 1, 11, tzinfo=timezone.utc)

def record(symbol="BTCUSDT", direction="LONG", price=100):
    return EntryLedgerRecord("l1", "e1", symbol, direction, "ENTRY", Decimal(str(price)), Decimal("2"), 1, "ISOLATED", Decimal("200"), Decimal("200"), Decimal("0.08"), Decimal("1"), Decimal("0"), Decimal("0"), Decimal("-1"), {"id": "snap"}, NOW, NOW, NOW, NOW)

def exit_record(symbol, direction, entry_price, exit_price):
    gross = (Decimal(str(exit_price)) - Decimal(str(entry_price))) * Decimal("2")
    if direction == "SHORT": gross = -gross
    return ExitLedgerRecord("x-" + symbol, "l1", "e-" + symbol, symbol, direction, "EXIT", "TAKE_PROFIT", Decimal(str(entry_price)), Decimal(str(exit_price)), Decimal("2"), 1, gross, Decimal("0.08"), Decimal("1"), Decimal("0"), gross - Decimal("1.08"), {"id": "snap"}, NOW, NOW, NOW, NOW)

def test_report_reads_ledger_and_filters():
    rows = [record(), record("ETHUSDT", "SHORT", 200)]
    assert len(build_report(rows)["trade_history"]) == 2
    assert len(build_report(rows, ReportFilters(symbol="ETHUSDT", direction="SHORT"))["trade_history"]) == 1

def test_openpyxl_export_sheets_turkish_and_log():
    target = ROOT / "reports" / "test_report.xlsx"
    log = export_xlsx([record()], target)
    workbook = load_workbook(target)
    assert tuple(workbook.sheetnames) == SHEETS
    assert log.library == "openpyxl" and log.library_version == "3.1.5"
    assert log.export_status == "PASS"

def test_missing_ledger_warning_and_live_lock():
    log = export_xlsx([], ROOT / "reports" / "empty.xlsx")
    assert log.export_status == "WARNING" and log.blocked_reason == "LEDGER_DATA_MISSING"
    config = json.loads((ROOT / "config/live_lock_config.json").read_text(encoding="utf-8"))
    assert config["LIVE_TRADING"] is False

def test_pnl_mismatch_is_blocking():
    row = ExitLedgerRecord("x1", "e1", "e2", "BTCUSDT", "LONG", "EXIT", "STOP_LOSS", Decimal("100"), Decimal("90"), Decimal("2"), 1, Decimal("999"), Decimal("0"), Decimal("0"), Decimal("0"), Decimal("999"), {"id": "snap"}, NOW, NOW, NOW, NOW)
    log = export_xlsx([row], ROOT / "reports" / "bad_report.xlsx")
    assert log.export_status == "BLOCKING_ERROR"
    assert log.blocked_reason == "LEDGER_PNL_MISMATCH"

def test_non_empty_long_short_fixture_export():
    entry_long = record("BTCUSDT", "LONG", 100)
    entry_short = record("ETHUSDT", "SHORT", 200)
    target = ROOT / "reports" / "sample_ledger_fixture.xlsx"
    fixture = [entry_long, exit_record("BTCUSDT", "LONG", 100, 110), entry_short, exit_record("ETHUSDT", "SHORT", 200, 190)]
    log = export_xlsx(fixture, target)
    workbook = load_workbook(target, read_only=True)
    assert log.export_status == "PASS"
    assert workbook["İşlem Geçmişi"].max_row == 5
    assert workbook["Config Snapshot"].max_row == 5
